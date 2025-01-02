import pandas as pd
import oracledb
oracledb.init_oracle_client(lib_dir=r"C:\Projetos_Python\BibliotecasOraclePython\instantclient_21_13")
import FuncoesAuxiliares
import pyautogui
from pywinauto import Application
import RepositorioDAO
import os
from pywinauto import Desktop
import csv
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

if __name__  == '__main__':
    print('EXECUTANDO TESTE02 DE JUNCAO_ARQUIVOS')

    t = (datetime.now() - relativedelta(months=1)).strftime("%m-%Y")
    print(t)

    #pyautogui.sleep(100000)

    varFuncao = FuncoesAuxiliares.Funcao_Apoio() 
    varExecuteDAO = RepositorioDAO.DAO()
    caminho =f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA1\\09-2024\\entradaLoja1.txt"
    pasta_ndd ="\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA1\\09-2024\\"
    arquivoNDD ="arquivoNDDLoja1.xlsx"

    entrada = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA3\\10-2024\\entradaLoja3.txt"
    arquivoNDD = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA3\\10-2024\\arquivoNDD_Loja3.csv"

    resultadoEntradas = varExecuteDAO.NF_pendentes_Ato_da_Entrega("27", "01-11-2024", "30-12-2024")
    
    # Carrega o arquivo XLSX
    workbook = load_workbook(filename="\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA27\\11-2024\\resultado_confronto27.xlsx")
    # Seleciona a primeira planilha
    sheet = workbook.active
    
    # Cria uma lista para armazenar os dados
    dados = []
    
    # Itera pelas linhas da planilha
    for row in sheet.iter_rows(values_only=True):
        # Adiciona cada linha como uma lista à lista de dados
        dados.append(list(row))
      

    resultadoJuncaoFinal = varFuncao.verificar_e_incluir(dados, resultadoEntradas)   


   # for valor in resultadoJuncaoFinal:
 #       print(f"Valor :{valor}")
    numeroLoja ="27"
    mes_ano="11-2024"

  #  resultado_formatado = [linha.split(";") for linha in resultadoJuncaoFinal]

    # Criar um DataFrame a partir da lista
    df = pd.DataFrame(resultadoJuncaoFinal, columns=["FORNECEDOR", "NUMERO", "CNPJ", "CHAVE", "EMISSÃO", "VALOR CONECT", "CFOP", "STATUS", "OPERAÇÃO"])

    # Definir o caminho do arquivo xlsx
    caminho_arquivo = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA{numeroLoja}\\{mes_ano}\\resultado_confronto_adicional{numeroLoja}.xlsx"

    # Salvar o DataFrame em um arquivo Excel
    df.to_excel(caminho_arquivo, index=False, engine='openpyxl')

    print(f"Arquivo salvo com sucesso em {caminho_arquivo}")






    pyautogui.sleep(100000)
    


    resultado = varFuncao.confronto_NDD(arquivoNDD, entrada)

    resultado_formatado = [linha.split(";") for linha in resultado]

    # Criar um DataFrame a partir da lista
    df = pd.DataFrame(resultado_formatado, columns=["FORNECEDOR", "NUMERO", "CNPJ", "CHAVE", "EMISSÃO", "VALOR CONECT", "CFOP", "STATUS"])

    # Definir o caminho do arquivo xlsx
    caminho_arquivo = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA3\\10-2024\\resultado_confronto3.xlsx"

    # Salvar o DataFrame em um arquivo Excel
    df.to_excel(caminho_arquivo, index=False, engine='openpyxl')

    print(f"Arquivo salvo com sucesso em {caminho_arquivo}")
   

    pyautogui.sleep(100000)
   
    resultado_formatado = [linha.split(";") for linha in resultado]

    #Salvar o DataFrame em um arquivo CSV
    caminho_arquivo = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA3\\10-2024\\resultado_confronto.csv"
    # Gravando a lista no arquivo CSV
    with open(caminho_arquivo, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        
        #Escrever o cabeçalho no CSV
        writer.writerow(["FORNECEDOR ;NUMERO ; CNPJ; CHAVE; EMISSÃO; STATUS"])


        # Escrever uma linha no CSV (a lista como linha única)
        writer.writerows(resultado_formatado)

    print(f"Arquivo salvo com sucesso em {caminho_arquivo}")

    """
    print("Retorno")
    
    # Iterar sobre os resultados e concatenar as informações
    for item_coluna_12, item_coluna_14 in zip(resultado_coluna_12, resultado_itens_presentes):
        # Concatenar as informações
        concatenado = f"Coluna 12: {item_coluna_12} - Itens Presentes: {item_coluna_14}"
        
        # Exibir na tela
        print(concatenado)

    print("Saiu do for")

   
    # Criar um DataFrame com os dados
    df_resultado = pd.DataFrame({
        "Coluna 12": resultado_coluna_12,
        "Itens Presentes": resultado_itens_presentes
    })

    # Salvar o DataFrame em um arquivo CSV
    caminho_arquivo = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA1\\09-2024\\resultado_confronto.csv"
    df_resultado.to_csv(caminho_arquivo, sep=';', index=False, encoding='ISO-8859-1')

    print(f"Dados salvos com sucesso em: {caminho_arquivo}")
    """
   
    """
    try:
        notas=[]
        resultado_coluna_12, resultado_itens_presentes = varFuncao.confronto_NDD("1")
        caminho_arquivo = f"\\\\10.11.10.3\\arcomixfs$\\Dados_Contabilidade\\FISCAL\\CONFRONTO_SPED\\LOJA1\\09-2024\\resultado_confronto.csv"

        # Iterar sobre os resultados e concatenar as informações
        for item_coluna_12, item_coluna_14 in zip(resultado_coluna_12, resultado_itens_presentes):
            concatenado = f"Coluna 12: {item_coluna_12} - Itens Presentes: {item_coluna_14}"
            notas.append(concatenado)
          #  print(concatenado)

        print("Saiu do for")

       # Criar um DataFrame
        df = pd.DataFrame(notas, columns=["Coluna"])

        # Nome do arquivo
        nome_arquivo = caminho_arquivo

        # Salvar no arquivo CSV
        df.to_csv(caminho_arquivo, sep=";", index=False, encoding="ISO-8859-1")

        print(f"Arquivo '{caminho_arquivo}' criado com sucesso.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
    """



    pyautogui.sleep(1000000)

    dtMesAnterior = varExecuteDAO.executaQuery("SELECT TO_CHAR(TRUNC(SYSDATE, 'MM') - INTERVAL '1' MONTH, 'DD/MM/YYYY') AS primeiro_dia_mes_anterior FROM dual")[0][0] 
    dt_Atual= varExecuteDAO.executaQuery("SELECT TO_CHAR(SYSDATE, 'DD/MM/YYYY') AS data_dia_Atual FROM dual")[0][0]  

    appSalvarArquivo = Application().connect(title_re=".*Caminho e nome.*")
    windowSalvar = appSalvarArquivo.Dialog
    windowSalvar.Wait('ready')
    comboboxArquivo = windowSalvar.ComboBox2
    pyautogui.sleep(1)
    comboboxArquivo.ClickInput()
    pyautogui.sleep(1)
    windowSalvar[u'ComboBox2'].type_keys(caminho) 
    pyautogui.sleep(2)
    windowSalvar[u'Sa&lvar'].click_input() #[u'Button', u'Sa&lvar', u'Button1', u'Button0', u'Sa&lvarButton']
    
    pyautogui.sleep(1000000)

    dao = RepositorioDAO.DAO()
    varFuncao = FuncoesAuxiliares.Funcao_Apoio()
    strLoja = None
    varQueryLojas = "select nroempresa, empresa from CONSINCO.dim_empresa where nroempresa not in (99, 800, 986, 987, 989, 999, 10, 13, 20, 25, 29)  order by NROEMPRESA"
    varLojas = dao.executaQuery(varQueryLojas)

    

    #LOOP DAS LOJAS
    for row in varLojas:
        lj=row[0]
        strLoja = str(lj)
      #  print(strLoja);
        dtaMovimento ="02/10/2024"
        validacaoxlsx =f'C:/Projetos_Python/TESOURARIA/arquivos/download/movimentolj{strLoja}.xlsx'

        if not os.path.exists(validacaoxlsx):
            continue

        
        varFuncao.xlsx_to_csv(strLoja)
        pyautogui.sleep(3)
        varFuncao.GeraSeqTurnoCSV(strLoja, dtaMovimento)


    """
    guptamdiframeAcer_Detalhamento = Application().connect(title_re=".*Movimento Detalhado.*")
    guptadialog = guptamdiframeAcer_Detalhamento[u'Gupta:Dialog']
    guptadialog.Wait('ready')
    guptachildtable = guptadialog[u'Gupta:ChildTable']
    guptachildtable.click_input() # Clica na tabela

    varFuncao = FuncoesAuxiliares.Funcao_Apoio()
       
    pyautogui.sleep(0.35)
    pyautogui.press('ctrl')
    pyautogui.sleep(0.35)
    pyautogui.press('insert') # Abre a linha para digitação 
    pyautogui.sleep(0.35)
    pyautogui.press('insert') # Inseri a linha 
 
    pyautogui.press('insert') # Abre a linha 
    pyautogui.sleep(0.35)
    pyautogui.write("787 - ")
    pyautogui.press('down')
    pyautogui.sleep(0.35)
    pyautogui.press('up')
    pyautogui.sleep(0.35)
    pyautogui.press('tab')
    pyautogui.sleep(0.35)
    pyautogui.write("5,00") # Informando o  valor 
    
    guptadialog[u'Button3'].click_input() #Click na no  botão conciliar
    pyautogui.sleep(2)

    guptadialog[u'Button6'].click_input() # Confirma e fecha a janela

    jnErro = varFuncao.check_window_exists("Movimento Detalhado")
    if jnErro == True:
        appDlgConf_Exclusao = Application().connect(title_re=".*Movimento Detalhado.*")
        window = appDlgConf_Exclusao.Dialog
        window.Wait('ready')
        button = window[u'Button6']
        button.click_input()
        pyautogui.sleep(1.5)



  

    


    df_csv = pd.read_csv('C:\\Projetos_Python\\TESOURARIA\\arquivos\\digitacao\\movimentolj3.csv')
        
    sql_query =
        SELECT h.seqturno, TO_CHAR(h.dtamovimento, 'DD/MM/YYYY') as Data, h.coo, h.nroempresa || h.nrocheckout || h.coo as KEY
        FROM consincomonitor.tb_docto h
        WHERE h.nroempresa = 3 
        AND h.dtamovimento = TO_DATE(SYSDATE - 20, 'DD/MM/YY')
        
          

    conection = oracledb.connect(user="consinco", password="consinco", dsn="10.102.227.2/arcomix.subnetarcomixda.vcnrootskyoneda.oraclevcn.com")
    df_sql = pd.read_sql_query(sql_query,conection)

    conection.close()

    print(df_csv.dtypes)
    print(df_sql.dtypes)

    #print(df_csv.columns)
    #print(df_sql.columns)

   # df_merged = pd.concat([df_csv,df_sql], axis=1)
   # print(df_merged)
   # df_merged.to_csv('C:\\Projetos_Python\\TESOURARIA\\arquivos\\digitacao\\arquivo_resultante.csv', index=False)

   ###########################################################################################################################################
    df_csv['KEY'] = df_csv['KEY'].astype(str)
    df_sql['KEY'] = df_sql['KEY'].astype(str)

    df_csv['KEY'] = df_csv['KEY'].astype(str)
    df_sql['KEY'] = df_sql['KEY'].astype(str)


    df_merged = pd.merge(df_csv, df_sql, on=['KEY','KEY'], how='inner')
    print(df_merged)
    df_merged.to_csv('C:\\Projetos_Python\\TESOURARIA\\arquivos\\digitacao\\arquivo_resultante.csv', index=True)

    """
